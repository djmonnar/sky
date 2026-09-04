"""
원가재고관리 엑셀 → 레시피 JSON (src/data/seed/haneulttang-recipes.json)

  python3 scripts/import-recipes-xlsx.py <원가재고관리.xlsx>

Sheet1: A 재료명 / B 구매가(1팩) / C 규격(kg 또는 L) / D kg(L)당 단가.
        B·C·D 가 빈 줄이 레시피 머리다. F 가 숫자면 사용량, "1회NNN" 이면 1회 제공 원가.
Sheet2: A 레시피 / B 재료 / C 수량(글) — 김치·소스 배치 레시피. 단가는 Sheet1 에서 찾는다.

숫자로 못 옮긴 수량은 note 에 원문을 남기고 수량을 0 으로 둔다 — 지어내지 않는다.
"""
import json, re, sys, unicodedata
import openpyxl

SRC = sys.argv[1]
OUT = "src/data/seed/haneulttang-recipes.json"

LIQUIDS = ("콜라", "탄산수", "사이다", "참기름", "간장", "쯔유", "식초", "빙초산", "액젓", "액기스",
           "냉면육수", "보리차", "물", "드레싱", "육수")
MENU = {"돼지갈비", "소갈비", "불고기", "된장찌개", "냉모밀", "비빔모밀"}
CATEGORY = {
    "돼지갈비": "구이", "소갈비": "구이", "불고기": "구이", "된장찌개": "식사", "냉모밀": "면", "비빔모밀": "면",
    "불고기 육수": "육수·소스", "파채소스": "육수·소스", "깻잎소스": "육수·소스", "쌈장": "육수·소스",
    "김반찬": "반찬", "쌈·쌈장": "반찬", "구이야채": "반찬", "겉절이": "반찬", "땡초찌": "반찬",
    "깻잎절임": "반찬", "백김치": "김치", "깍두기": "김치", "샐러드": "반찬", "테이블 기본제공": "기타",
}

def clean(s):
    return re.sub(r"\s+", " ", unicodedata.normalize("NFC", str(s or ""))).strip()

def num(v):
    return float(v) if isinstance(v, (int, float)) else None

def base_name(name):
    """규격이 붙은 이름에서 재료 이름만: '탄산수 1.5L' → '탄산수', '해찬들 된장 14KG' → '해찬들 된장'"""
    n = re.sub(r"\s*\d+(\.\d+)?\s*(kg|KG|Kg|l|L|B|봉|g)\b", "", name).strip()
    return n or name

EA_ITEMS = ("모밀면",)   # 1팩 단위로 사는 것 — 규격 1 이 kg 이 아니라 1묶음이다

def unit_for(name, pack):
    n = name.replace(" ", "")
    if any(k in n for k in EA_ITEMS): return "ea"
    if any(k in n for k in LIQUIDS): return "L"
    return "kg"

wb = openpyxl.load_workbook(SRC, data_only=True)
s1, s2 = wb.worksheets[0], wb.worksheets[1]

# ---------- Sheet1: 재료 단가표 + 레시피 묶음 ----------
recipes = []      # [{name, ingredients:[...], flags}]
prices = {}       # base name → {unit, unitCost, packPrice, packSize}
current = None
ing_seq = 0

def add_price(name, b, c, d):
    key = base_name(name).replace(" ", "")
    if d and d > 0 and key not in prices:
        prices[key] = {"unit": unit_for(name, c), "unitCost": round(d, 2), "packPrice": b, "packSize": c}

def new_recipe(name):
    r = {"name": name, "ingredients": [], "needs": []}
    recipes.append(r)
    return r

for row in s1.iter_rows(min_row=1, max_row=140, max_col=8, values_only=True):
    a, b, c, d, e, f, g, h = row
    name = clean(a)
    bn, cn, dn = num(b), num(c), num(d)
    if not name and not any(v is not None for v in row):
        continue
    # 머리줄: 이름만 있고 값이 없다. '#DIV/0!' 은 단가를 못 셈한 재료지 머리가 아니다.
    header_text = isinstance(d, str) and d != "#DIV/0!"   # 첫 줄 '돼지갈비 | 단가 | 규격 | KG당단가'
    is_header = bool(name) and (bn is None or bn == 0) and cn is None and (d in (None, 0) or header_text)
    if not name and dn == 0:
        current = new_recipe("땡초찌");  # 겉절이 뒤 빈 머리 — 다음 줄 땡초찌
        continue
    if is_header:
        title = {"쌈,쌈장": "쌈·쌈장", "된장찌개": "된장찌개", "비빔모밀": "비빔모밀", "깻잎절임": "깻잎절임", "구이야채": "구이야채"}.get(name, name)
        if title in ("땡초찌",) and current and current["name"] == "땡초찌":
            pass
        else:
            current = new_recipe(title)
        continue
    if current is None:
        continue
    if name in ("보리차원액", "물", "숯") and current["name"] in ("샐러드",):
        current = new_recipe("테이블 기본제공")
    if not name:
        continue
    add_price(name, bn, cn, dn)
    if dn is None or dn == 0 or not isinstance(d, (int, float)):
        # 단가를 못 셈한 줄(#DIV/0!, 빈칸) — 이름만 남긴다
        current["ingredients"].append({"name": base_name(name), "quantity": 0, "unit": "kg", "unitCost": 0, "note": "단가 없음"})
        current["needs"].append(base_name(name))
        continue
    unit = unit_for(name, cn)
    unit_cost = round(dn, 2)
    qty, note = 0.0, None
    fn = num(f)
    ftxt = clean(f) if isinstance(f, str) else ""
    per_serving = None
    if fn is not None and fn > 0:
        # 20 이상이면 g 로 적은 것이다 (498.6kg 꽃게는 없다)
        qty = fn / 1000 if fn >= 20 else fn
        note = f"엑셀 사용량 {f}{'g' if fn >= 20 else unit}"
    elif ftxt:
        m = re.search(r"(\d[\d,]*(?:\.\d+)?)", ftxt.replace("1회", " ").replace("1장당", " ").replace("제공", " "))
        if m:
            per_serving = float(m.group(1).replace(",", ""))
            note = f"엑셀 1회 제공 원가 {ftxt}"
    if per_serving is None and isinstance(f, str) and num(g):
        per_serving = num(g); note = f"엑셀 1회 제공 원가 {g}원"
    if per_serving is not None and unit_cost > 0:
        qty = round(per_serving / unit_cost, 4)
    if qty == 0:
        note = "수량 미입력 — 엑셀에 사용량 없음"
        current["needs"].append(base_name(name))
    current["ingredients"].append({"name": base_name(name), "quantity": round(qty, 4), "unit": unit, "unitCost": unit_cost, **({"note": note} if note else {})})

# ---------- Sheet2: 배치 레시피 (수량 있음, 단가는 단가표에서) ----------
ALIAS = {"물(풀용)": None, "물": None, "천일염(무절임용)": "소금", "설탕(무절임용)": "설탕", "천일염": "소금", "무(채썰기)": "무",
         "당근(채썰기)": "당근", "황물엿": "물엿", "두배사과식초": "2배사과식초", "재래식콩된장": "해찬들 된장", "대파": "대파"}
def lookup(name):
    key = (ALIAS.get(name, name) or "").replace(" ", "")
    if not key: return None
    for k, v in prices.items():
        if k == key or key in k or k in key:
            return v
    return None

QTY = re.compile(r"^(\d+(?:\.\d+)?)\s*(kg|g|l|L|ml|개|박스|box|Box|B)$")
batches = {}
cur = None
for a, b, c, d in s2.iter_rows(min_row=2, max_col=4, values_only=True):
    a, b, c = clean(a), clean(b), clean(c)
    if a and b == "재료명":
        cur = a; batches[cur] = []; continue
    if a: cur = a
    if cur is None or not b: continue
    batches.setdefault(cur, []).append((b, c))

sheet2_recipes = []
for title, rows in batches.items():
    name = re.sub(r"\(.*?\)", "", title).strip()
    servings_note = title[len(name):].strip("() ")
    r = {"name": name, "ingredients": [], "needs": [], "memo": f"엑셀 배치 기준{f' {servings_note}' if servings_note else ''}"}
    for ing, qtext in rows:
        m = QTY.match(qtext.replace(" ", ""))
        price = lookup(ing)
        unit_cost = price["unitCost"] if price else 0
        base_unit = price["unit"] if price else "kg"
        if m:
            q, u = float(m.group(1)), m.group(2)
            u = {"l": "L", "개": "ea", "box": "박스", "Box": "박스", "B": "박스"}.get(u, u)
            # 단가는 kg/L 당이다 — g/ml 로 적힌 수량은 그 단위에 맞춘 단가로
            uc = unit_cost
            if u == "g" and base_unit == "kg": uc = unit_cost / 1000
            elif u == "ml" and base_unit == "L": uc = unit_cost / 1000
            item = {"name": ing, "quantity": q, "unit": u, "unitCost": round(uc, 4)}
            if u in ("ea", "박스") and price and base_unit in ("kg", "L"):
                # kg 단가로 산 것을 '개'로 적었다 — 개당 무게를 모르니 단가를 지어내지 않는다
                item["unitCost"] = 0
                item["note"] = f"{base_unit} 단가 {unit_cost:,.0f}원 — 개당 무게로 환산 필요"
                r["needs"].append(ing)
            elif item["unitCost"] == 0:
                r["needs"].append(ing); item["note"] = "단가 없음"
        else:
            item = {"name": ing, "quantity": 0, "unit": base_unit, "unitCost": round(unit_cost, 2), "note": f"엑셀 수량: {qtext}"}
            r["needs"].append(ing)
        r["ingredients"].append(item)
    sheet2_recipes.append(r)

# 깍두기·백김치는 Sheet2 에 수량이 있으니 그것으로 바꿔 끼운다
by_name = {r["name"]: r for r in recipes}
for r in sheet2_recipes:
    if r["name"] in by_name:
        by_name[r["name"]]["ingredients"] = r["ingredients"]
        by_name[r["name"]]["needs"] = r["needs"]
        by_name[r["name"]]["memo"] = r["memo"]
    else:
        recipes.append(r)

out = []
for i, r in enumerate(recipes, 1):
    kind = "menu" if r["name"] in MENU else "side"
    ings = []
    for j, ing in enumerate(r["ingredients"], 1):
        ings.append({"id": f"r{i}-{j}", **ing})
    needs = sorted(set(r["needs"]))
    memo_parts = [r.get("memo", "엑셀 원가재고관리에서 옮김")]
    if kind == "menu": memo_parts.append("판매가·기준 인분을 채워 주세요.")
    if needs: memo_parts.append("수량·단가 확인 필요: " + ", ".join(needs))
    out.append({
        "id": i, "name": r["name"], "category": CATEGORY.get(r["name"], "기타"), "kind": kind,
        "servings": 1, "salePrice": 0, "ingredients": ings, "memo": " ".join(memo_parts), "active": True,
    })

json.dump(out, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
print(f"{len(out)} recipes → {OUT}")
for r in out:
    cost = sum(i["quantity"] * i["unitCost"] for i in r["ingredients"])
    print(f"  [{r['kind']:4}] {r['name']:10} 재료 {len(r['ingredients']):2}개  원가 {cost:>10,.0f}원  {r['memo'][:60]}")
